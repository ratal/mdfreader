# -*- coding: utf-8 -*-
""" Measured Data Format file reader main module

Platform and python version
----------------------------------------
With Unix and Windows for python 2.6+ and 3.2+

:Author: `Aymeric Rateau <http://code.google.com/p/mdfreader/>`__

Created on Sun Oct 10 12:57:28 2010

Dependencies
-------------------
- Python >2.6, >3.2 <http://www.python.org>
- Numpy >1.6 <http://numpy.scipy.org>
- Sympy to convert channels with formula
- Matplotlib >1.0 <http://matplotlib.sourceforge.net>
- NetCDF
- h5py for the HDF5 export
- xlwt for the excel export (not existing for python3)
- openpyxl for the excel 2007 export
- scipy for the Matlab file conversion

Attributes
--------------
PythonVersion : float
    Python version currently running, needed for compatibility of both python 2.6+ and 3.2+
    
mdfreader module
--------------------------
"""
from io import open
from struct import unpack
from math import ceil
from mdf3reader import mdf3
from mdf4reader import mdf4
from numpy import arange, interp, all, diff, mean, vstack,hstack, float64, zeros, empty, delete
from numpy import nan, datetime64, array

from argparse import ArgumentParser
from sys import version_info
from os.path import splitext
PythonVersion=version_info
PythonVersion=PythonVersion[0]

def convertMatlabName(channel):
    """Removes non allowed characters for a Matlab variable name
    
    Parameters
    -----------------
    channel : string
        channel name
        
    Returns
    -----------
    string
        channel name compatible for Matlab
    """
    if PythonVersion<3:
        channel=channel.decode('utf-8')
    channelName=channel.replace('[', '_ls_')
    channelName=channelName.replace(']', '_rs_')
    channelName=channelName.replace('$', '')
    channelName=channelName.replace('.', 'p')
    channelName=channelName.replace('\\','_bs_')
    channelName=channelName.replace('/','_fs_')
    channelName=channelName.replace('(','_lp_')
    channelName=channelName.replace(',','_rp_')
    channelName=channelName.replace('@','_am_')
    channelName=channelName.replace(' ','_')
    channelName=channelName.replace(':','_co_')
    channelName=channelName.replace('-','_hy_')
    channelName=channelName.replace('-','_hy_')
    return channelName

class mdfinfo( dict):
    """MDFINFO is a class gathering information from block headers in a MDF (Measure Data Format) file
        Structure: nested dicts. Primary key is Block type, then data group, channel group and channel number.
        Examples of dicts:
    - mdfinfo['HDBlock'] header block
    - mdfinfo['DGBlock'][dataGroup] Data Group block
    - mdfinfo['CGBlock'][dataGroup][channelGroup] Channel Group block
    - mdfinfo['CNBlock'][dataGroup][channelGroup][channel] Channel block including text blocks for comment and identifier
    - mdfinfo['CCBlock'][dataGroup][channelGroup][channel] Channel conversion information
    
    Attributes
    --------------
    fileName : str
        file name
    mdfversion : int
        mdf file version number
        
    Methods
    ------------
    readinfo( fileName = None, filterChannelNames=False )
        Reads MDF file and extracts its complete structure
    listChannels( fileName = None )
        Read MDF file blocks and returns a list of contained channels
    
    Examples
    --------------
    >>> import mdfreader
    >>> FILENAME='toto.dat'
    >>> yop=mdfreader.mdfinfo(FILENAME)
    or if you are just interested to have only list of channels
    >>> yop=mdfreader.mdfinfo() # creates new instance f mdfinfo class
    >>> yop=mdfreader.listChannels(FILENAME) # returns a simple list of channel names
    """

    def __init__( self, fileName = None, filterChannelNames = False ):
        """ You can give optionally to constructor a file name that will be parsed
        
        Parameters
        ----------------
        fileName : str, optional
            file name
        filterChannelNames : bool, optional
            flag to filter long channel names including module names separated by a '.'
        """
        self.fileName = fileName
        self.mdfversion = 0
        if fileName != None:
            self.readinfo( fileName, filterChannelNames )

    def readinfo( self, fileName = None, filterChannelNames=False ):
        """ Reads MDF file and extracts its complete structure
        
        Parameters
        ----------------
        fileName : str, optional
            file name. If not input, uses fileName attribute
        filterChannelNames : bool, optional
            flag to filter long channel names including module names separated by a '.'
        """
        if self.fileName == None:
            self.fileName = fileName
        # Open file
        try:
            fid = open( self.fileName, 'rb' )
        except IOError:
            print('Can not find file'+self.fileName)
            raise
        # read Identifier block
        fid.seek(28)
        VersionNumber = unpack( '<H', fid.read( 2 ) )
        self.mdfversion = VersionNumber[0]
        if self.mdfversion < 400: # up to version 3.x not compatible with version 4.x
            from mdfinfo3 import info3
            self.update(info3(None, fid, filterChannelNames))
        else: #MDF version 4.x
            from mdfinfo4 import info4
            self.update(info4(None, fid))

    def listChannels( self, fileName = None ):
        """ Read MDF file blocks and returns a list of contained channels
        
        Parameters
        ----------------
        fileName : string
            file name
            
        Returns
        -----------
        nameList : list of string
            list of channel names
        """
        if self.fileName == None:
            self.fileName = fileName
        # Open file
        try:
            fid = open( self.fileName, 'rb' )
        except IOError:
            print('Can not find file'+self.fileName)
            raise
        # read Identifier block
        fid.seek(28)
        VersionNumber=unpack( '<H', fid.read( 2 ) )
        self.mdfversion = VersionNumber[0]
        if self.mdfversion < 400: # up to version 3.x not compatible with version 4.x
            from mdfinfo3 import info3
            channelNameList=info3()
            nameList=channelNameList.listChannels3(self.fileName)
        else:
            from mdfinfo4 import info4
            channelNameList=info4()
            nameList=channelNameList.listChannels4(self.fileName)
        return nameList

class mdf( mdf3,  mdf4 ):
    """ mdf class
    
    Attributes
    --------------
    fileName : str
        file name
    VersionNumber : int
        mdf file version number
    masterChannelList : dict
        Represents data structure: a key per master channel with corresponding value containing a list of channels
        One key or master channel represents then a data group having same sampling interval.
    multiProc : bool
        Flag to request channel conversion multi processed for performance improvement.
        One thread per data group.
    author : str
    organisation : str
    project : str
    subject : str
    comment : str
    time : str
    date : str
    
    Methods
    ------------
    read( fileName = None, multiProc = False, channelList=None, convertAfterRead=True, filterChannelNames=False )
        reads mdf file version 3.x and 4.x
    write( fileName=None )
        writes simple mdf 3.3 file
    getChannelData( channelName )
        returns channel numpy array
    convertAllChannel()
        converts all channel data according to CCBlock information
    getChannelUnit( channelName )
        returns channel unit
    plot( channels )
        Plot channels with Matplotlib
    resample( samplingTime = 0.1, masterChannel=None )
        Resamples all data groups
    exportToCSV( filename = None, sampling = 0.1 )
        Exports mdf data into CSV file
    exportToNetCDF( filename = None, sampling = None )
        Exports mdf data into netcdf file
    exportToHDF5( filename = None, sampling = None )
        Exports mdf class data structure into hdf5 file
    exportToMatlab( filename = None )
        Exports mdf class data structure into Matlab file
    exportToExcel( filename = None )
        Exports mdf data into excel 95 to 2003 file
    exportToXlsx( filename=None )
        Exports mdf data into excel 2007 and 2010 file
    convertToPandas( sampling=None )
        converts mdf data structure into pandas dataframe(s)
    keepChannels( channelList )
        keeps only list of channels and removes the other channels
    mergeMdf( mdfClass ):
        Merges data of 2 mdf classes
    copy()
        copy a mdf class
    
    Notes
    --------
    mdf class is a nested dict
    Channel name is the primary dict key of mdf class
    At a higher level, each channel includes the following keys :
        - 'data' : containing vector of data (numpy)
        - 'unit' : unit (string)
        - 'master' : master channel of channel (time, crank angle, etc.)
        - 'description' : Description of channel
        - 'conversion': mdfinfo nested dict for CCBlock.
            Exist if channel not converted, used to convert with getChannelData method
    
    Examples
    --------------
    >>> import mdfreader
    >>> yop=mdfreader.mdf('NameOfFile')
    >>> yop.keys() # list channels names
    >>> yop.masterChannelList() # list channels grouped by raster or master channel
    >>> yop.plot('channelName') or yop.plot({'channel1','channel2'})
    >>> yop.resample(0.1) or yop.resample(channelName='master3')
    >>> yop.exportoCSV(sampling=0.01)
    >>> yop.exportNetCDF()
    >>> yop.exporttoHDF5()
    >>> yop.exporttoMatlab()
    >>> yop.exporttoExcel()
    >>> yop.exporttoXlsx()
    >>> yop.convertToPandas() # converts data groups into pandas dataframes
    >>> yop.keepChannels({'channel1','channel2','channel3'}) # drops all the channels except the one in argument
    >>> yop.getChannelData('channelName') # returns channel numpy array
    """
    
    def __init__( self, fileName = None,  channelList=None, convertAfterRead=True, filterChannelNames=False):
        """ mdf class constructor.
        When mdf class is constructed, constructor can be called to directly reads file
        
        Parameters
        ----------------
        fileName : str, optional
            file name
        
        channelList : list of str, optional
            list of channel names to be read
            If you use channelList, reading might be much slower but it will save you memory. Can be used to read big files
        
        convertAfterRead : bool, optional
            flag to convert channel after read, True by default
            If you use convertAfterRead by setting it to false, all data from channels will be kept raw, no conversion applied.
            If many float are stored in file, you can gain from 3 to 4 times memory footprint
            To calculate value from channel, you can then use method .getChannelData()
            
        filterChannelNames : bool, optional
            flag to filter long channel names from its module names separated by '.'
        """
        self.fileName = None
        self.VersionNumber=None
        self.masterChannelList = {}
        self.author=''
        self.organisation=''
        self.project=''
        self.subject=''
        self.comment=''
        self.time=''
        self.date=''
        self.multiProc = False # flag to control multiprocessing, default deactivate, giving priority to mdfconverter
        self.convert_tables = False # if True converts raw data with expensive loops, not necessary most cases
        # clears class from previous reading and avoid to mess up
        self.clear()
        if not fileName == None:
            self.read( fileName, channelList=channelList, convertAfterRead=convertAfterRead, filterChannelNames=filterChannelNames )
            self.fileName=fileName

    def read( self, fileName = None, multiProc = False, channelList=None, convertAfterRead=True, filterChannelNames=False):
        """ reads mdf file version 3.x and 4.x
        
        Parameters
        ----------------
        fileName : str, optional
            file name
        
        multiProc : bool
            flag to activate multiprocessing of channel data conversion
        
        channelList : list of str, optional
            list of channel names to be read
            If you use channelList, reading might be much slower but it will save you memory. Can be used to read big files
        
        convertAfterRead : bool, optional
            flag to convert channel after read, True by default
            If you use convertAfterRead by setting it to false, all data from channels will be kept raw, no conversion applied.
            If many float are stored in file, you can gain from 3 to 4 times memory footprint
            To calculate value from channel, you can then use method .getChannelData()
            
        filterChannelNames : bool, optional
            flag to filter long channel names from its module names separated by '.'
        
        Notes
        --------
        If you keep convertAfterRead to true, you can set attribute mdf.multiProc to activate channel conversion in multiprocessing.
        Gain in reading time can be around 30% if file is big and using a lot of float channels
        
        Warning:
        ------------
        MultiProc use should be avoided when reading several files in a batch, it is not thread safe.
        You should better multi process instances of mdf rather than using multiproc in mdf class (see implementation of mdfconverter)
        """
        if self.fileName == None:
            self.fileName = fileName
        
        print(self.fileName)
        
        # read file blocks
        info=mdfinfo(self.fileName, filterChannelNames)
        
        self.VersionNumber=info.mdfversion
        if self.VersionNumber<400: # up to version 3.x not compatible with version 4.x
            self.read3(self.fileName, info, multiProc, channelList, convertAfterRead)
        else: #MDF version 4.x. Channel by channel reading implemented
            self.read4(self.fileName, info, multiProc, channelList, convertAfterRead)
    
    def write(self, fileName=None):
        """Writes simple mdf 3.3 file
        
        Parameters
        ----------------
        fileName : str, optional
            Name of file
            If file name is not input, written file name will be the one read with appended '_new' string before extension
        
        Notes
        --------
        All channels will be converted, so size might be bigger than original file
        """
        if fileName is None:
            splitName=splitext(self.fileName)
            self.fileName=splitName[-2]+'_New'+splitName[-1]
        else:
            self.fileName=fileName
        # makes sure all channels are converted
        self.convertAllChannel()
        self.write3(fileName=self.fileName)

    def getChannelData(self, channelName):
        """Return channel numpy array
        
        Parameters
        ----------------
        channelName : str
            channel name
            
        Returns:
        -----------
        numpy array
            converted, if not already done, data corresponding to channel name
        
        Notes
        ------
        This method is the safest to get channel data as numpy array from 'data' dict key might contain raw data
        """
        if self.VersionNumber<400:
            return self.getChannelData3(channelName)
        else:
            return self.getChannelData4(channelName)
    
    def convertAllChannel(self):
        """Converts all channels from raw data to converted data according to CCBlock information
        Converted data will take more memory.
        """
        if self.VersionNumber<400:
            return self.convertAllChannel3()
        else:
            return self.convertAllChannel4()
    
    def getChannelUnit(self, channelName):
        """Returns channel unit string
        Implemented for a future integration of pint
        
        Parameters
        ----------------
        channelName : str
            channel name
            
        Returns
        -----------
        str
            unit string description
        """
        return self[channelName]['unit']

    def plot( self, channels ):
        """Plot channels with Matplotlib
        
        Parameters
        ----------------
        channels : str or list of str
            channel name or list of channel names
            
        Notes
        ---------
        Channel description and unit will be tentatively displayed with axis labels
        """
        try:
            import matplotlib.pyplot as plt
        except:
            print('matplotlib not found' )
            raise
        if type(channels) is str:
            channels={channels}
        for channelName in channels:
            if channelName in self:
                data = self.getChannelData(channelName)
                if not data.dtype.kind in ['S', 'U']: # if channel not a string
                    self.fig = plt.figure()
                    # plot using matplotlib the channel versus master channel
                    if len(list(self.masterChannelList.keys()))==1: # Resampled signals
                        masterName = list(self.masterChannelList.keys())[0]
                        if not masterName: # resampled channels, only one time channel most probably called 'master'
                            masterName ='master'
                        if masterName in list(self.keys()): # time channel properly defined
                            plt.plot( self.getChannelData(masterName), data)
                            plt.xlabel( masterName + ' [' + self.getChannelUnit(masterName) + ']' )
                        else: # no time channel found
                            plt.plot( data )
                    else: # not resampled
                        if self[channelName]['master'] in list(self.keys()): # master channel is proper channel name
                            plt.plot( self.getChannelData(self[channelName]['master']), data )
                            plt.xlabel( self[channelName]['master'] + ' [' + self.getChannelUnit(self[channelName]['master']) + ']' )
                        else:
                            plt.plot( data )

                    plt.title( self[channelName]['description'])
                    if self.getChannelUnit(channelName) == {}:
                        plt.ylabel( channelName )
                    else:
                        plt.ylabel( channelName + ' [' + self.getChannelUnit(channelName) + ']' )
                    plt.grid( True )
                    plt.show()
            else:
                print(( 'Channel ' + channelName + ' not existing' ))

    def allPlot( self ):
        # plot all channels in the object, be careful for test purpose only,
        # can display many many many plots overloading your computer
        for Name in list(self.keys()):
            try:
                self.plot( Name )
            except:
                print( Name )

    def resample( self, samplingTime = 0.1, masterChannel=None ):
        """ Resamples all data groups into one data group having defined
        sampling interval or sharing same master channel
        
        Parameters
        ----------------
        samplingTime : float
            resampling interval
        **or**
        masterChannel : str
            master channel name used for all channels
        
        Notes
        --------
        1. resampling is relatively safe for mdf3 as it contains only time series.
        However, mdf4 can contain also distance, angle, etc. It might make not sense 
        to apply one resampling to several data groups that do not share same kind 
        of master channel (like time resampling to distance or angle data groups)
        If several kind of data groups are used, you should better use pandas to resample
        
        2. resampling will convert all your channels so be careful for big files
        and memory consumption
        """
        # must make sure all channels are converted
        self.convertAllChannel()
        # resample all channels to one sampling time vector
        if len(list(self.masterChannelList.keys()))>1: # Not yet resampled
            channelNames = list(self.keys())
            minTime = maxTime = []
            if masterChannel is None: # create master channel if not proposed
                masterChannelName='master'
                self[masterChannelName] = {}
                unit = ''
                masterType = 1 # time by default
                
                for master in list(self.masterChannelList.keys()):
                    masterData = self.getChannelData(master)
                    if master in self and len( masterData ) > 5: # consider groups having minimum size 
                        minTime.append( masterData[0] )
                        maxTime.append( masterData[len( masterData ) - 1] )
                        if len(self.getChannelUnit(master))>1 :
                            unit = self.getChannelUnit(master)
                            masterType = self[master]['masterType']
                self[masterChannelName]['data'] = arange( min( minTime ),max( maxTime ),samplingTime )
                self[masterChannelName]['unit'] = unit
                self[masterChannelName]['description'] = 'Unique master channel'
                self[masterChannelName]['masterType'] = masterType
            else:
                masterChannelName=masterChannel

            # Interpolate channels
            timevect=[]
            for Name in channelNames:
                try:
                    if Name not in list(self.masterChannelList.keys()): # not a master channel
                        timevect = self.getChannelData(self[Name]['master'])
                        if not self.getChannelData(Name).dtype.kind in ('S', 'U'): # if channel not array of string
                            self[Name]['data'] = interp( self.getChannelData(masterChannelName), timevect, self.getChannelData(Name) )
                            if masterChannelName in self[Name]:
                                del self[Name][masterChannelName]
                        else: # can not interpolate strings, remove channel containing string
                            self.masterChannelList[self[Name]['master']].remove(Name)
                            self.pop(Name)
                except:
                    if len( timevect ) != len( self.getChannelData(Name) ):
                        print(( Name + ' and time channel ' + self[Name][masterChannelName] + ' do not have same length' ))
                    elif not all( diff( timevect ) > 0 ):
                        print(( Name + ' has non regularly increasing time channel ' + self[Name][masterChannelName] ))
            # remove time channels in masterChannelList
            for ind in list(self.masterChannelList.keys()):
                del self[ind]
            self.masterChannelList = {} # empty dict
            self.masterChannelList[masterChannelName] = list(self.keys())
        else:
            print('Already resampled')

    def exportToCSV( self, filename = None, sampling = 0.1 ):
        """Exports mdf data into CSV file
        
        Parameters
        ----------------
        filename : str, optional
            file name. If no name defined, it will use original mdf name and path
        
        sampling : float, optional
            sampling interval. By default, sampling is 0.1sec but can be changed
            
        Notes
        --------
        Data saved in CSV fille be automatically resampled as it is difficult to save in this format
        data not sharing same master channel
        Warning: this can be slow for big data, CSV is text format after all
        """
        import csv
        self.resample( sampling )
        if filename == None:
            filename = splitext(self.fileName)[0]
            filename = filename + '.csv'
        if PythonVersion <3:
            f = open( filename, "wb")
        else:
            f = open( filename, "wt" , encoding='latin-1')
        writer = csv.writer( f, dialect = csv.excel )
        # writes header
        writer.writerow( [name for name in list(self.keys()) if self.getChannelData(name).dtype in ('float64','float32')] ) # writes channel names
        writer.writerow( [(self.getChannelUnit(name)) for name in list(self.keys()) if self.getChannelData(name).dtype in ('float64','float32')] ) # writes units
        # concatenate all channels
        buf = vstack( [self.getChannelData(name).transpose() for name in list(self.keys()) if self.getChannelData(name).dtype in ('float64','float32')] )
        buf = buf.transpose()
        # Write all rows
        r, c = buf.shape
        writer.writerows( [list( buf[i, :] ) for i in range( r )] )
        f.close()

    def exportToNetCDF( self, filename = None, sampling = None ):
        """Exports mdf data into netcdf file
        
        Parameters
        ----------------
        filename : str, optional
            file name. If no name defined, it will use original mdf name and path
        
        sampling : float, optional
            sampling interval.
            
        Dependency
        -----------------
        scipy
        """
        try:
            from scipy.io import netcdf
        except:
            print( 'scipy.io module not found' )
            raise
        def cleanName( name ):
            allowedStr=' ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-+_.@'
            buf=''
            for c in name:
                if c in allowedStr:
                    buf+=c
            return buf

        def setAttribute(f, name, value):
            if len(value)>0: # netcdf does not allow empty strings...
                setattr( f, name, value)
            else:
                pass
        if sampling != None:
            self.resample( sampling )
        if filename == None:
            filename = splitext(self.fileName)[0]
            filename = filename + '.nc'
        f = netcdf.netcdf_file( filename, 'w' )
        setAttribute( f, 'Date', (self.date))
        setAttribute( f, 'Time', (self.time))
        if self.author is not None:
            setAttribute(f, 'Author', self.author)
        if self.organisation is not None:
            setAttribute( f, 'Organization', (self.organisation))
        if self.project is not None:
            setAttribute( f, 'ProjectName', (self.project))
        if self.subject is not None:
            setAttribute( f, 'Subject', (self.subject))
        if self.comment is not None:
            setAttribute( f, 'Comment', (self.comment))
        # Create dimensions having name of all time channels
        for time in list(self.masterChannelList.keys()):
            f.createDimension( time, len( self.getChannelData(time) ) )
        # Create variables definition, dimension and attributes
        var = {}
        for name in list(self.keys()):
            data=self.getChannelData(name)
            if data.dtype == 'float64':
                type = 'd'
            elif data.dtype == 'float32':
                type = 'f'
            elif data.dtype in ['int8', 'int16', 'uint8', 'uint16']:
                type = 'h'
            elif data.dtype in ['int32', 'uint32']:
                type = 'i'
            elif data.dtype.kind in ['S', 'U'] :
                type = 'c'
            else:
                print(( 'Can not process numpy type ' + str(data.dtype) + ' of channel' ))
            # create variable
            CleanedName = cleanName( name )
            if len( list(self.masterChannelList.keys()) ) == 1: # mdf resampled
                var[name] = f.createVariable( CleanedName, type, ( list(self.masterChannelList.keys())[0], ) )
            else: # not resampled
                var[name] = f.createVariable( CleanedName, type, ( self[name]['master'], ) )
            # Create attributes
            setAttribute( var[name], 'title', CleanedName )
            setAttribute( var[name], 'units', cleanName(self.getChannelUnit(name)))
            setAttribute( var[name], 'Description', cleanName(self[name]['description']))
            if name in list(self.masterChannelList.keys()):
                setAttribute( var[name], 'Type', 'Master Channel' )
                setAttribute( var[name], 'datatype', 'master' )
            else:
                setAttribute( var[name], 'Type', 'Data Channel' )
        # put data in variables
        for name in list(self.keys()):
            var[name] = self.getChannelData(name)
        f.close()

    def exportToHDF5( self, filename = None, sampling = None ):
        """Exports mdf class data structure into hdf5 file
        
        Parameters
        ----------------
        filename : str, optional
            file name. If no name defined, it will use original mdf name and path
            
        sampling : float, optional
            sampling interval.
            
        Dependency
        ------------------
        h5py
        
        Notes
        --------
        The maximum attributes will be stored
        Data structure will be similar has it is in masterChannelList attribute
        """
        # 
        try:
            import h5py
            import os
        except:
            print( 'h5py not found' )
            raise
        if sampling != None:
            self.resample( sampling )
        if filename == None:
            filename = splitext(self.fileName)[0]
            filename = filename + '.hdf'
        f = h5py.File( filename, 'w' ) # create hdf5 file
        filegroup=f.create_group(os.path.basename(filename)) # create group in root associated to file
        filegroup.attrs['Author']=self.author
        filegroup.attrs['Date']=self.date
        filegroup.attrs['Time']= self.time 
        filegroup.attrs['Organization']=self.organisation
        filegroup.attrs['ProjectName']=self.project
        filegroup.attrs['Subject']=self.subject
        filegroup.attrs['Comment']=self.comment
        if len( list(self.masterChannelList.keys()) ) > 1:
            # if several time groups of channels, not resampled
            groups = {}
            ngroups = 0
            grp = {}
            for channel in list(self.keys()):
                if self[channel]['master'] not in list(groups.keys()):
                    # create new time group
                    ngroups += 1
                    groups[self[channel]['master'] ] = ngroups
                    grp[ngroups] = filegroup.create_group( self[channel]['master'] )
                dset = grp[groups[self[channel]['master'] ]].create_dataset( channel, data = self.getChannelData(channel) )
                dset.attrs[ 'unit']=self.getChannelUnit(channel)
                dset.attrs['description']=self[channel]['description']
        else: # resampled or only one time for all channels : no groups
            for channel in list(self.keys()):
                channelName=convertMatlabName(channel)
                dset = filegroup.create_dataset( channelName, data = self.getChannelData(channel) )
                dset.attrs[ 'unit']=self.getChannelUnit(channel)
                dset.attrs['description']=self[channel]['description']
        f.close()

    def exportToMatlab( self, filename = None ):
        """Export mdf data into Matlab file format 5, tentatively compressed
        
        Parameters
        ----------------
        filename : str, optional
            file name. If no name defined, it will use original mdf name and path
            
        Dependency
        ------------------
        scipy
            
        Notes
        --------
        This method will dump all data into Matlab file but you will loose below information:
        - unit and descriptions of channel
        - data structure, what is corresponding master channel to a channel. Channels might have then different lengths
        """
        # export class data struture into .mat file
        try:
            from scipy.io import savemat
        except:
            print( 'scipy module not found' )
            raise
        if filename == None:
            filename = splitext(self.fileName)[0]
            filename = filename + '.mat'
        # convert self into simple dict without and metadata
        temp = {}
        for channel in list(self.keys()):
            data=self.getChannelData(channel)
            if not data.dtype.kind in ('S', 'U'): # does not like special characters chains, skip
                channelName=convertMatlabName(channel)
                temp[channelName] = data
        try: # depends of version used , compression can be used
            savemat( filename , temp, long_field_names = True,format='5',do_compression=True,oned_as='column' )
        except:
            savemat( filename , temp, long_field_names = True,format='5')

    def exportToExcel( self , filename = None ):
        """Exports mdf data into excel 95 to 2003 file
        
        Parameters
        ----------------
        filename : str, optional
            file name. If no name defined, it will use original mdf name and path
            
        Dependencies
        --------------------
        xlwt for python 2.6+
        xlwt3 for python 3.2+
            
        Notes
        --------
        xlwt is not fast for even for small files, consider other binary formats like HDF5 or Matlab
        If there are more than 256 channels, data will be saved over different worksheets
        Also Excel 203 is becoming rare these days
        """
        try:
            if PythonVersion<3:
                import xlwt
            else:
                import xlwt3 as xlwt
        except:
            print( 'xlwt module missing' )
            raise
        if filename == None:
            filename = filename = splitext(self.fileName)[0]
            filename = filename + '.xls'
        styleText = xlwt.easyxf( 'font: name Times New Roman, color-index black, bold off' )
        coding='utf-8'
        wb = xlwt.Workbook(encoding=coding)
        channelList = list(self.keys())
        if PythonVersion<3:
            Units=[ self.getChannelUnit(channel).decode(coding, 'replace') for channel in list(self.keys())]
        else:
            Units=[ self.getChannelUnit(channel) for channel in list(self.keys())]
        # Excel 2003 limits
        maxCols = 255
        maxLines = 65535
        workbooknumber = int( ceil( len( channelList ) * 1.0 / ( maxCols * 1.0 ) ) )
        tooLongChannels = []
        # split colmuns in several worksheets if more than 256 cols
        for workbook in range( workbooknumber ):
            ws = wb.add_sheet( 'Sheet' + str( workbook ) ) #, cell_overwrite_ok = True )
            if workbook == workbooknumber - 1: # last sheet
                columnrange = list(range( workbook * maxCols, len( channelList )))
            elif workbook < workbooknumber - 1 and workbooknumber > 1: # first sheets
                columnrange = list(range( workbook * maxCols, ( workbook + 1 ) * maxCols))
            for col in columnrange:
                # write header
                ws.write( 0, col - workbook * maxCols, channelList[col] , styleText )
                ws.write( 1, col - workbook * maxCols, Units[col] , styleText )
                vect = self.getChannelData(channelList[col]) # data vector
                if not len( vect ) > maxLines :
                    if  vect.dtype.kind not in ['S', 'U']: # if not a string or unicode
                        [ws.row( row + 2 ).set_cell_number( col - workbook * maxCols, vect[row] ) for row in list(range( len( vect ) ))]
                    else: # it's a string, cannot write for the moment
                        if PythonVersion <3:
                            try:
                                vect=vect.encode(coding)
                            except:
                                pass
                        [ws.row( row + 2 ).set_cell_text( col - workbook * maxCols, vect[row]) for row in list(range( len( vect ) ))]
                else: # channel too long, written until max Excel line limit
                    if vect.dtype.kind not in ['S', 'U']: # if not a string
                        [ws.row( row + 2 ).set_cell_number( col - workbook * maxCols, vect[row] ) for row in list(range( maxLines ))]
                    else: # it's a string, cannot write for the moment
                        if PythonVersion <3:
                            vect=vect.encode(coding)
                        [ws.row( row + 2 ).set_cell_text( col - workbook * maxCols, vect[row] ) for row in list(range( maxLines ))]
                    tooLongChannels.append( channelList[col] ) # to later warn user the channel is not completely written
        wb.save( filename ) # writes workbook on HDD
        if len( tooLongChannels ) > 0: # if not empty, some channels have been not processed
            print( 'Following channels were too long to be processed completely, maybe you should resample : ' )
            print( tooLongChannels )

    def exportToXlsx(self, filename=None):
        """Exports mdf data into excel 2007 and 2010 file
        
        Parameters
        ----------------
        filename : str, optional
            file name. If no name defined, it will use original mdf name and path
            
        Dependency
        -----------------
        openpyxl
            
        Notes
        --------
        It is recommended to export resampled data for performances
        """
        try:
            import openpyxl
        except:
            print('Module openpyxl missing')
            raise
        if filename == None:
            filename = splitext(self.fileName)[0]
            filename = filename + '.xlsx'
        channels=list(self.keys())
        maxRows=max([len(self.getChannelData(channel)) for channel in list(self.keys())]) # find max column length
        maxCols=len(list(self.keys())) # number of columns
        print('Creating Excel sheet')
        if len( list(self.masterChannelList.keys()) ) > 1: # not resampled data, can be long, writing cell by cell !
            wb=openpyxl.workbook.Workbook(encoding='utf-8')
            ws=wb.get_active_sheet()
            # write header
            if PythonVersion<3:
                for j in range(maxCols):
                    ws.cell(row=0, column=j).value=channels[j].decode('utf-8', 'ignore')
                    ws.cell(row=1, column=j).value=self.getChannelUnit(channels[j]).decode('utf-8', 'ignore')
            else:
                for j in range(maxCols):
                    ws.cell(row=0, column=j).value=channels[j]
                    ws.cell(row=1, column=j).value=self.getChannelUnit(channels[j])            
            for j in range(maxCols):
                print(channels[j])
                data = self.getChannelData(channels[j])
                if data.dtype  in ['int8', 'int16', 'uint8', 'uint16']:
                     for r in range(len(data)):
                        ws.cell(row=r+2, column=j).value=float64(data[r])
                else:
                     for r in range(len(data)):
                        ws.cell(row=r+2, column=j).value=data[r]
        else: # resampled data
            wb=openpyxl.workbook.Workbook(optimized_write=True, encoding='utf-8')
            ws=wb.create_sheet()
            # write header
            ws.append(channels)
            ws.append([ self.getChannelUnit(channel) for channel in list(self.keys())])
            # write data
            maxRows=max([len(self.getChannelData(channel)) for channel in list(self.keys())]) # find max column length
            maxCols=len(list(self.keys())) # number of columns
            bigmat=zeros(maxRows) # create empty column
            buf=bigmat
            for col in range(maxCols):
                data = self.getChannelData(channels[col])
                if not data.dtype.kind in ['S', 'U']:
                    chanlen=len(data)
                    if chanlen<maxRows:
                        buf[:]=None
                        buf[0:chanlen]=data
                        bigmat=vstack((bigmat, buf))
                    else:
                        bigmat=vstack((bigmat, data))
                else:
                    buf[:]=None
                    bigmat=vstack((bigmat, buf))
            bigmat=delete(bigmat, 0, 0)
            [ws.append(bigmat[:, row]) for row in range(maxRows)]
        print('Writing file, please wait')
        wb.save(filename)

    def keepChannels(self, channelList):
        """ keeps only list of channels and removes the other channels
        
        Parameters
        ----------------
        channelList : list of str
            list of channel names
        """
        channelList=[channel for channel in channelList]
        removeChannels=[]
        for channel in list(self.keys()):
            if channel not in channelList and not 'master' in channel and channel not in list(self.masterChannelList.keys()) :
                # avoid to remove master channels otherwise problems with resample
                removeChannels.append(channel)
        if not len(removeChannels)==0:
            [self.masterChannelList[self[channel]['master']].remove(channel) for channel in removeChannels]
            [self.pop(channel) for channel in removeChannels]

    def copy(self):
        """copy a mdf class
            
        Returns:
        ------------
        mdf class instance
            copy of a mdf class
        """
        yop=mdf()
        yop.multiProc=self.multiProc
        yop.fileName=self.fileName
        yop.masterChannelList=self.masterChannelList
        for channel in list(self.keys()):
            yop[channel]=self[channel]
        return yop

    def mergeMdf(self, mdfClass):
        """Merges data of 2 mdf classes
        
        Parameters
        ----------------
        mdfClass : mdf
            mdf class instance to be merge with self
            
        Notes
        --------
        both classes must have been resampled, otherwise, impossible to know master channel to match
        create union of both channel lists and fill with Nan for unknown sections in channels
        """
        self.convertAllChannel() # make sure all channels are converted
        unionedList=list(mdfClass.keys()) and list(self.keys())
        if not len(list(self.masterChannelList.keys()))==1:
            print( 'Data not resampled')
            raise
        initialTimeSize=len(self.getChannelData('master'))
        for channel in unionedList:
            data = self.getChannelData(channel)
            mdfData = mdfClass.getChannelData(channel)
            if channel in mdfClass and channel in self: # channel exists in both class
                if not channel=='master':
                    self[channel]['data']=hstack((data, mdfData))
                else:
                    offset=mean(diff(mdfData)) # sampling
                    offset=data[-1]+offset # time offset
                    self[channel]['data']=hstack((data, mdfData+offset))
            elif channel in mdfClass: # new channel for self from mdfClass
                self[channel]=mdfClass[channel] # initialise all fields, units, descriptions, etc.
                refill=empty(initialTimeSize)
                refill.fil(nan) # fill with NANs
                self[channel]['data']=hstack((refill,  mdfData)) # readjust against time
            else: #channel missing in mdfClass
                refill=empty(len(mdfClass.getChannelData('master')))
                refill.fill(nan) # fill with NANs
                self[channel]['data']=hstack((data, refill))

    def convertToPandas(self, sampling=None):
        """converts mdf data structure into pandas dataframe(s)
        
        Parameters
        ----------------
        sampling : float, optional
            resampling interval
            
        Notes
        --------
        One pandas dataframe is converted per data group
        Not adapted yet for mdf4 as it considers only time master channels
        """
        # convert data structure into pandas module
        try:
            import pandas as pd
        except:
            print('Module pandas missing')
            raise
        if sampling is not None:
            self.resample(sampling)
        datetimeInfo=datetime64(self.date.replace(':','-')+'T'+self.time)
        originalKeys=list(self.keys())
        for group in list(self.masterChannelList.keys()):
            temp={}
            # convert time channel into timedelta
            time=datetimeInfo+array(self.getChannelData(group)*10E6, dtype='timedelta64[us]')
            for channel in self.masterChannelList[group]:
                temp[channel]=pd.Series(self.getChannelData(channel), index=time)
            self[group+'_group']=pd.DataFrame(temp)
            self[group+'_group'].pop(group) # delete time channel, no need anymore
        # clean rest of self from data and time channel information
        [self[channel].pop('data') for channel in originalKeys]
        [self[channel].pop('master') for channel in originalKeys]
        self.masterGroups=[] # save time groups name in list
        [self.masterGroups.append(group+'_group') for group in list(self.masterChannelList.keys())]
        self.masterChannelList={}

if __name__ == "__main__":
    try:
        from multiprocessing import freeze_support
        freeze_support()
    except:
        None
    parser=ArgumentParser(prog='mdfreader', description='reads mdf file')
    parser.add_argument('--convertAfterRead', default=True, help = 'True by default, flag to convert raw channel data to physical values just after reading. Deactivate if you have memory concerns')
    parser.add_argument('--filterChannelNames', default=False,  help = 'False by default, activates channel name filtering, removes modules names separated by a point character')
    parser.add_argument('fileName',  help='mdf file name', required=True)
    parser.add_argumetn('--channelList', dest='channelList', type=list, default=None, help = 'list of channels to read')
    
    args = parser.parse_args()
    mdf(fileName=args.fileName, channelList=args.channelList, converAfterRead=args.convertAfterRead, filterChannelNames=args.filterChannelNames)
