# -*- coding: utf-8 -*-
""" Measured Data Format file reader main module

Platform and python version
----------------------------------------
With Unix and Windows for python 2.6+ and 3.2+

:Author: `Aymeric Rateau <https://github.com/ratal/mdfreader>`__

Created on Sun Oct 10 12:57:28 2010

Dependencies
-------------------
- Python >3.4 <http://www.python.org>
- Numpy >1.14 <http://numpy.scipy.org>
- Sympy to convert channels with formula
- bitarray for not byte aligned data parsing
- Matplotlib >1.0 <http://matplotlib.sourceforge.net>
- scipy for NetCDF
- h5py for the HDF5 export
- xlwt for the excel export (not existing for python3)
- openpyxl >2.0 for the excel 2007 export
- hdf5storage for the Matlab file conversion
- zlib to uncompress data block if needed

mdfreader
--------------------------
"""

from io import open
from struct import unpack
from math import ceil
from os.path import splitext
from os import remove
from os import name as osname
from warnings import warn
from datetime import datetime
from argparse import ArgumentParser
from numpy import arange, linspace, interp, all, diff, mean, vstack, hstack, float64, float32
from numpy import nan, datetime64, array, searchsorted, clip, empty
from numpy.ma import MaskedArray, masked, empty as ma_empty
from .mdf3reader import Mdf3
from .mdf4reader import Mdf4
from .mdf import _open_mdf, dataField, descriptionField, unitField, masterField, masterTypeField, idField
from .mdfinfo3 import Info3, _generate_dummy_mdf3
from .mdfinfo4 import Info4, _generate_dummy_mdf4


def _convert_to_matlab_name(channel):
    """Removes non allowed characters for a Matlab variable name

    Parameters
    -----------------
    channel : string
        channel name

    Returns
    -----------
    string
        channel name compatible for Matlab
    """
    channel_name = channel.replace('[', '_ls_')
    channel_name = channel_name.replace(']', '_rs_')
    channel_name = channel_name.replace('$', '')
    channel_name = channel_name.replace('.', 'p')
    channel_name = channel_name.replace('\\', '_bs_')
    channel_name = channel_name.replace('/', '_fs_')
    channel_name = channel_name.replace('(', '_lp_')
    channel_name = channel_name.replace(')', '_rp_')
    channel_name = channel_name.replace(',', '_c_')
    channel_name = channel_name.replace('@', '_am_')
    channel_name = channel_name.replace(' ', '_')
    channel_name = channel_name.replace(':', '_co_')
    channel_name = channel_name.replace('-', '_hy_')
    channel_name = channel_name.replace('-', '_hy_')

    def clean_name(name):
        allowed_str = 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789_.'
        buf = ''
        for c in name:
            if c in allowed_str:
                buf += c
        return buf

    channel_name = clean_name(channel_name)
    # limit the variable length at 63 character, Matlab limitation
    # if you use long names including modules names separated by a '.'
    # you can use filter_channel_names=True parameter
    channel_name = channel_name[:63]
    return channel_name


def _convert_to_hdf5_name(channel):
    """Removes non allowed characters for a hdf5 variable name

    Parameters
    -----------------
    channel : string
        channel name

    Returns
    -----------
    string
        channel name compatible for hdf5
    """
    channel_name = channel.replace('/', '_fs_')
    channel_name = channel_name.replace('\\', '_bs_')
    return channel_name


class MdfInfo(dict):
    __slots__ = ['fileName', 'fid', 'zipfile', 'mdfversion', 'filterChannelNames']
    """ MDFINFO is a class gathering information from block headers in a MDF (Measure Data Format) file.
    Structure is nested dicts. Primary key is Block type, then data group, channel group and channel number.
    Examples of dicts
    - mdfinfo['HDBlock'] header block
    - mdfinfo['DGBlock'][dataGroup] Data Group block
    - mdfinfo['CGBlock'][dataGroup][channelGroup] Channel Group block
    - mdfinfo['CNBlock'][dataGroup][channelGroup][channel] Channel block including text blocks for comment and identifier
    - mdfinfo['CCBlock'][dataGroup][channelGroup][channel] Channel conversion information

    Attributes
    --------------
    fileName : str
        file name
    mdfversion : int
        mdf file version number
    filterChannelNames : bool
        flag to filter long channel names including module names separated by a '.'
    fid
        file identifier
    zipfile
        flag to indicate the mdf4 is packaged in a zip

    Methods
    ------------
    read_info( fileName = None, filterChannelNames=False )
        Reads MDF file and extracts its complete structure
    list_channels( file_name = None )
        Read MDF file blocks and returns a list of contained channels

    Examples
    --------------
    >>> import mdfreader
    >>> FILENAME='toto.dat'
    >>> yop=mdfreader.MdfInfo(FILENAME)
    or if you are just interested to have only list of channels
    >>> yop=mdfreader.MdfInfo() # creates new instance of mdfinfo class
    >>> yop.list_channels(FILENAME) # returns a simple list of channel names
    """

    def __init__(self, file_name=None, filter_channel_names=False, fid=None, minimal=0):

        """ You can give optionally to constructor a file name that will be parsed

        Parameters
        ----------------
        file_name : str, optional
            file name
        filter_channel_names : bool, optional
            flag to filter long channel names including module names separated by a '.'
        fid : file identifier, optional
        """

        self.fileName = file_name
        self.filterChannelNames = filter_channel_names
        self.mdfversion = 410
        self.fid = fid
        self.zipfile = False
        if file_name is not None:
            self.read_info(file_name, fid, minimal)

    def read_info(self, file_name=None, fid=None, minimal=0):

        """ Reads MDF file and extracts its complete structure

        Parameters
        ----------------
        file_name : str, optional
            file name. If not input, uses fileName attribute
        fid : file identifier, optional
        minimal : int
            0 will load every metadata
            1 will load DG, CG, CN and CC
            2 will load only DG
        """

        if self.fileName is None or file_name is not None:
            self.fileName = file_name

        # Open file
        if self.fid is None or (self.fid is not None and self.fid.closed):
            (self.fid, self.fileName, self.zipfile) = _open_mdf(self.fileName)

        # read Identifier block
        self.fid.seek(28)
        mdf_version_number = unpack('<H', self.fid.read(2))
        self.mdfversion = mdf_version_number[0]
        if self.mdfversion < 400:  # up to version 3.x not compatible with version 4.x
            self.update(Info3(None, self.fid, self.filterChannelNames, minimal))
        else:  # MDF version 4.x
            self.update(Info4(None, self.fid, self.filterChannelNames, minimal))
            if self.zipfile and fid is None:  # not from mdfreader.read()
                remove(self.fileName)

    def list_channels(self, file_name=None):

        """ Read MDF file blocks and returns a list of contained channels

        Parameters
        ----------------
        file_name : string
            file name

        Returns
        -----------
        nameList : list of string
            list of channel names
        """

        if self.fileName is None or file_name is not None:
            self.fileName = file_name
        # Open file
        (self.fid, self.fileName, zipfile) = _open_mdf(self.fileName)
        # read Identifier block
        self.fid.seek(28)
        mdf_version_number = unpack('<H', self.fid.read(2))
        self.mdfversion = mdf_version_number[0]
        if self.mdfversion < 400:  # up to version 3.x not compatible with version 4.x
            channel_name_list = Info3()
            name_list = channel_name_list.list_channels3(self.fileName, self.fid)
        else:
            channel_name_list = Info4()
            name_list = channel_name_list.list_channels4(self.fileName, self.fid)
            if zipfile:  # not from mdfreader.read()
                remove(self.fileName)
        return name_list

    def list_channels_sources(self, file_name=None):

        """ Read MDF file blocks and returns a list of contained channels and
        source message information


        Parameters
        ----------------
        file_name : string
            file name

        Returns
        -----------
        nameList : list of string
            list of channel names with source message information
        """

        if self.fileName is None or file_name is not None:
            self.fileName = file_name
        # Open file
        (self.fid, self.fileName, zipfile) = _open_mdf(self.fileName)
        # read Identifier block
        self.fid.seek(28)
        mdf_version_number = unpack('<H', self.fid.read(2))
        self.mdfversion = mdf_version_number[0]
        if self.mdfversion >= 400:  # up to version 4.x not compatible with version 3.x
            channel_name_sources_list = Info4()
            name_sources_list = channel_name_sources_list.list_channels_sources4(self.fileName, self.fid)
            if zipfile:  # not from mdfreader.read()
                remove(self.fileName)
        return name_sources_list

    def _generate_dummy_mdf(self, channel_list=None):
        """ Parse MDF file structure and create a dummy mdf object structure

        Parameters
        ----------------
        channel_list : str, optional
            list of channels
        """
        if self.mdfversion < 400:  # up to version 3.x not compatible with version 4.x
            return _generate_dummy_mdf3(self, channel_list)
        else:  # MDF version 4.x
            return _generate_dummy_mdf4(self, channel_list)


class Mdf(Mdf4, Mdf3):
    """ Mdf class

    Attributes
    --------------
    fileName : str
        file name
    MDFVersionNumber : int
        mdf file version number
    masterChannelList : dict
        Represents data structure: a key per master channel with corresponding value containing a list of channels
        One key or master channel represents then a data group having same sampling interval.
    multiProc : bool
        Flag to request channel conversion multi processed for performance improvement.
        One thread per data group.
    fileMetadata : dict
        file metadata with minimum keys : author, organisation, project, subject, comment, time, date

    Methods
    ------------
    read( file_name = None, multi_processed = False, channel_list=None, convert_after_read=True,
    filter_channel_names=False, no_data_loading=False, compression=False)
        reads mdf file version 3.x and 4.x
    write( file_name=None )
        writes simple mdf file
    get_channel_data( channel_name )
        returns channel numpy array
    convert_all_channel()
        converts all channel data according to CCBlock information
    get_channel_unit( channel_name )
        returns channel unit
    plot( channels )
        Plot channels with Matplotlib
    resample( sampling_time = 0.1, master_channel=None )
        Resamples all data groups
    export_to_csv( file_name = None, sampling = 0.1 )
        Exports mdf data into CSV file
    export_to_NetCDF( file_name = None, sampling = None )
        Exports mdf data into netcdf file
    export_to_hdf5( file_name = None, sampling = None )
        Exports mdf class data structure into hdf5 file
    export_to_matlab( file_name = None )
        Exports mdf class data structure into Matlab file
    export_to_excel( file_name = None )
        Exports mdf data into excel 95 to 2003 file
    export_to_xlsx( file_name=None )
        Exports mdf data into excel 2007 and 2010 file
    convert_to_pandas( sampling=None )
        converts mdf data structure into pandas dataframe(s)
    keep_channels( channel_list )
        keeps only list of channels and removes the other channels
    merge_mdf( mdf_class ):
        Merges data of 2 mdf classes

    Notes
    --------
    mdf class is a nested dict.
    Channel name is the primary dict key of mdf class.
    At a higher level, each channel includes the following keys :
        - 'data' : containing vector of data (numpy)
        - 'unit' : unit (string)
        - 'master' : master channel of channel (time, crank angle, etc.)
        - 'description' : Description of channel
        - 'conversion': mdfinfo nested dict for CCBlock.
            Exist if channel not converted, used to convert with getChannelData method

    Examples
    --------------
    >>> import mdfreader
    >>> yop=mdfreader.Mdf('NameOfFile')
    >>> yop.keys() # list channels names
    # list channels grouped by raster or master channel
    >>> yop.masterChannelList
    >>> yop.plot('channelName') or yop.plot({'channel1','channel2'})
    >>> yop.resample(0.1) or yop.resample()
    >>> yop.export_to_csv(sampling=0.01)
    >>> yop.export_to_NetCDF()
    >>> yop.export_to_hdf5()
    >>> yop.export_to_matlab()
    >>> yop.export_to_excel()
    >>> yop.export_to_xlsx()
    >>> yop.convert_to_pandas() # converts data groups into pandas dataframes
    >>> yop.write() # writes mdf file
    # drops all the channels except the one in argument
    >>> yop.keep_channels(['channel1','channel2','channel3'])
    >>> yop.get_channel_data('channelName') # returns channel numpy array
    >>> yop=mdfreader.Mdf()  # create an empty Mdf object
    # add channel in Mdf object
    >>> yop.add_channel(channel_name, data, master_channel, master_type, unit='lumen', description='what you want')
    >>> yop.write('filename') # change version with yop.MDFVersionNumber or specifically use write3/4()
    """

    def read(self, file_name=None, multi_processed=False, channel_list=None, convert_after_read=True,
             filter_channel_names=False, no_data_loading=False, compression=False, metadata=2, source_list=None):
        """ reads mdf file version 3.x and 4.x

        Parameters
        ----------------
        file_name : str, optional
            file name

        multi_processed : bool
            flag to activate multiprocessing of channel data conversion.

        channel_list : list of str, optional
            list of channel names to be read.
            If you use channelList, reading might be much slower but it will save you memory.
            Can be used to read big files.

        convert_after_read : bool, optional
            flag to convert channel after read, True by default.
            If you use convertAfterRead by setting it to false, all data from channels will be kept raw,
            no conversion applied. If many float are stored in file, you can gain from 3 to 4 times memory footprint.
            To calculate value from channel, you can then use method getChannelData()

        filter_channel_names : bool, optional
            flag to filter long channel names from its module names separated by '.'

        no_data_loading : bool, optional
            Flag to read only file info but no data to have minimum memory use.

        compression : bool or str, optional
            To compress data in memory using blosc or bcolz, takes cpu time.
            if compression = int(1 to 9), uses bcolz for compression.
            if compression = 'blosc', uses blosc for compression.
            Choice given, efficiency depends of data.

        metadata: int, optional, default = 2
            Reading metadata has impact on performance, especially for mdf 4.x using xml.
            2: minimal metadata reading (mostly channel blocks).
            1: used for noDataLoading.
            0: all metadata reading, including Source Information, Attachment, etc..

        source_list: list, optional, default = None
            list containing the source messages to identify what device send the
            different message.

        Notes
        --------
        If you keep convertAfterRead to true, you can set attribute mdf.multiProc to activate channel conversion
         in multiprocessing. Gain in reading time can be around 30% if file is big and using a lot of float channels

        Warning
        ------------
        MultiProc use should be avoided when reading several files in a batch, it is not thread safe.
        You should better multi process instances of mdf rather than using multiproc in mdf class
        (see implementation of mdfconverter)
        """
        if self.fileName is None or file_name is not None:
            self.fileName = file_name

        # Open file
        (self.fid, self.fileName, self.zipfile) = _open_mdf(self.fileName)

        # read Identifier block
        self.fid.seek(28)
        mdf_version_number = unpack('<H', self.fid.read(2))
        self.MDFVersionNumber = mdf_version_number[0]

        if self.MDFVersionNumber < 400:  # up to version 3.x not compatible with version 4.x
            if not no_data_loading:
                self.read3(self.fileName, None, multi_processed, channel_list,
                           convert_after_read, filter_channel_names, compression)
            else:  # populate minimum mdf structure
                self._noDataLoading = True
                self.info = Info3(None, fid=self.fid,
                                  filter_channel_names=filter_channel_names, minimal=1)
                (self.masterChannelList, mdf_dict) = _generate_dummy_mdf3(self.info, channel_list)
                self.update(mdf_dict)
        else:  # MDF version 4.x
            if not no_data_loading:
                self.read4(self.fileName, None, multi_processed, channel_list,
                           convert_after_read, filter_channel_names, compression, metadata, source_list)
            else:  # populate minimum mdf structure
                self._noDataLoading = True
                self.info = Info4(None, fid=self.fid,
                                  filter_channel_names=filter_channel_names, minimal=1)
                (self.masterChannelList, mdf_dict) = _generate_dummy_mdf4(self.info, channel_list)
                self.update(mdf_dict)

        if not self.fid.closed:  # close file
            self.fid.close()

    def write(self, file_name=None, compression=False, column_oriented=False):
        """Writes simple mdf file, same format as originally read, default is 4.x

        Parameters
        ----------------
        file_name : str, optional
            Name of file
            If file name is not input, written file name will be the one read with
            appended '_new' string before extension
        compression : bool
            Flag to store data compressed (from mdf version 4.1)
            If activated, will write in version 4.1 even if original file is in version 3.x
        column_oriented : bool
            Flag to store , column oriented channels

        Notes
        --------
        All channels will be converted, so size might be bigger than original file
        """
        if file_name is None:
            split_name = splitext(self.fileName)
            if split_name[-1] in ('.mfxz', '.MFXZ'):
                split_name[-1] = '.mfx'  # do not resave in compressed file
            file_name = ''.join([split_name[-2], '_New', split_name[-1]])
        # makes sure all channels are converted
        self.convert_all_channels()
        if self.MDFVersionNumber < 400 and not compression:
            self.write3(file_name=file_name)
        else:
            self.write4(file_name=file_name, compression=compression, column_oriented=column_oriented)

    def get_channel_data(self, channel_name, raw_data=False):
        """Return channel numpy array

        Parameters
        ----------------
        channel_name : str
            channel name
        raw_data: bool
            flag to return non converted data

        Returns
        -----------
        numpy array
            converted, if not already done, data corresponding to channel name

        Notes
        ------
        This method is the safest to get channel data as numpy array from 'data' dict key might contain raw data
        """
        if self.MDFVersionNumber < 400:
            vector = self._get_channel_data3(channel_name, raw_data)
        else:
            vector = self._get_channel_data4(channel_name, raw_data)
        if self._noDataLoading:
            # remove data loaded in object to save memory
            self.set_channel_data(channel_name, None)
        return vector

    def convert_all_channels(self):
        """Converts all channels from raw data to converted data according to CCBlock information.
        Converted data will take more memory.
        """
        if self.MDFVersionNumber < 400:
            return self._convert_all_channel3()
        else:
            return self._convert_all_channel4()

    def plot(self, channel_name_list_of_list):
        """Plot channels with Matplotlib

        Parameters
        ----------------
        channel_name_list_of_list : str or list of str or list of list of str
            channel name or list of channel names or list of list of channel names
            list of list will create multiplots

        Notes
        ---------
        Channel description and unit will be tentatively displayed with axis labels
        """
        try:
            import matplotlib.pyplot as plt
        except ImportError:
            warn('matplotlib not found')
            return
        try:
            from mpldatacursor import datacursor
            cursor_possible = True
        except ImportError:
            warn('no cursor available, to use cursor, please install mpldatacursor')
            cursor_possible = False
        if isinstance(channel_name_list_of_list, str):
            channel_name_list_of_list = [channel_name_list_of_list]  # converts in list
        for channel_name_list in channel_name_list_of_list:
            fig = plt.subplot(len(channel_name_list_of_list), 1, channel_name_list_of_list.index(channel_name_list) + 1)
            if isinstance(channel_name_list, str):
                channel_name_list = [channel_name_list]  # converts in list
            for channelName in channel_name_list:
                if channelName in self:
                    data = self.get_channel_data(channelName)
                    if data.dtype.kind not in ['S', 'U']:  # if channel not a string
                        # self.fig = plt.figure()  # could be needed
                        # plot using matplotlib the channel versus master channel
                        if len(self.masterChannelList) == 1:  # Resampled signals or only one master
                            master_name = list(self.masterChannelList)[0]
                            if not master_name:  # resampled channels, only one time channel probably called 'master'
                                master_name = 'master'
                            master_data = self.get_channel_data(master_name)
                            if master_data is None:  # no master channel
                                master_data = arange(0, len(data), 1)
                            if master_name in self.masterChannelList:  # time channel properly defined
                                plt.plot(master_data, data, label=channelName)
                                plt.xlabel('{0} [{1}]'.format(master_name, self.get_channel_unit(master_name)))
                            else:  # no time channel found
                                plt.plot(data)
                        else:  # not resampled
                            master_name = self.get_channel_master(channelName)
                            master_data = self.get_channel_data(master_name)
                            if master_data is None:  # no master channel
                                master_data = arange(0, len(data), 1)
                            if master_name in self.masterChannelList:  # master channel is proper channel name
                                plt.plot(master_data, data, label=channelName)
                                plt.xlabel('{0} [{1}]'.format(master_name, self.get_channel_unit(master_name)))
                            else:
                                plt.plot(data)

                        plt.title(self.get_channel_desc(channelName))
                        if self.get_channel_unit(channelName) == {}:
                            plt.ylabel(channelName)
                        else:
                            plt.ylabel('{0} [{1}]'.format(channelName, self.get_channel_unit(channelName)))
                else:
                    warn('Channel {} not existing'.format(channelName))
            plt.grid(True)
            plt.legend(loc="upper left", frameon=False)
            plt.title('')
        if cursor_possible:
            datacursor()
        plt.show()
        return fig

    def plot_all(self):
        # plot all channels in the object, be careful for test purpose only,
        # can display many many many plots overloading your computer
        for Name in self:
            try:
                self.plot(Name)
            except:
                warn(Name)

    def resample(self, sampling=None, channel=None, master_channel=None):
        """ Resamples as much as possible all data groups into one data group having defined
        sampling interval or sharing same defined master channel

        Parameters
        ----------------
        sampling : float, optional
            resampling interval, None by default. If None, will rely on channel or master_channel
             parameters to define reference data group. If both are undefined, picking the first master
        ** or | and **
        channel : str, optional
            channel name to be resampled
        ** or | and **
        master_channel : str, optional
            master channel name to be used as reference

        Notes
        --------
        1. resampling will be applied only to master channels that have same type as the one
        given by channel or master_channel parameters (applicable only to mdf4)

        2. resampling will convert all your channels so be careful for big files
        and memory consumption
        """
        if self:  # mdf contains data
            # must make sure all channels are converted
            self.convert_all_channels()

            if channel is not None and master_channel is None:
                master_channel_name = self.get_channel_master(channel)
                master_channel_type = self.get_channel_master_type(master_channel_name)
            elif channel is None and master_channel is None:
                # pick the first master
                master_channel_name = list(self.masterChannelList.keys())[0]
                master_channel_type = self.get_channel_master_type(master_channel_name)
            else:
                master_channel_name = master_channel
                master_channel_type = self.get_channel_master_type(master_channel)
            master_data = self.get_channel_data(master_channel_name)
            if sampling is not None:
                master_data = arange(master_data[0], master_data[-1], sampling)

            if master_channel_name is None or \
                    master_channel_name not in self.masterChannelList[master_channel_name]:
                # No master channel in selected group, looking at other groups
                min_master = []
                max_master = []
                length = []
                master_channel_name = 'master'
                for master in self.masterChannelList:
                    if master is not None and master != '' and \
                            master in self and self.masterChannelList[master] and \
                            self.get_channel_master_type(master) == master_channel_type:
                        master_data = self.get_channel_data(master)
                        # consider groups having minimum size
                        if master in self and len(master_data) > 5:
                            min_master.append(master_data[0])
                            max_master.append(master_data[-1])
                            length.append(len(master_data))
                            master_channel_name = master
                if min_master:  # at least 1 datagroup has a master channel to be resampled
                    if sampling is None:
                        master_data = linspace(min(min_master), max(max_master), num=max(length))
                    else:
                        master_data = arange(min(min_master), max(max_master), sampling)
                    self.add_channel(master_channel_name, master_data, master_channel_name,
                                     master_type=self.get_channel_master_type(master),
                                     unit=self.get_channel_unit(master),
                                     description=self.get_channel_desc(master), conversion=None)
                else:
                    warn('no master channel existing, considering first channel as master !')
                    current_master = list(self.masterChannelList.keys())[0]
                    # pick the first channel from the first data group to be master, maybe lucky
                    master_channel_name = self.masterChannelList[current_master][0]
                    # changing master to first channel
                    self.masterChannelList[master_channel_name] = self.masterChannelList.pop(current_master)
                    for channel in self:
                        self.set_channel_master(channel, master_channel_name)
                    master_channel_type = self.get_channel_master_type(master_channel_name)
                    master_data = self.get_channel_data(master_channel_name)
                    if sampling is not None:
                        master_data = arange(master_data[0], master_data[-1], sampling)

            # Interpolate channels
            for master in list(self.masterChannelList.keys()):
                if self.get_channel_master_type(master) == master_channel_type:
                    self.resample_group(sampling, master, new_master_data=master_data)
                    # remove old master channel
                    if not master_channel_name == master:
                        self.masterChannelList[master].remove(master)  # removing previous master from list
                        for channel in self.masterChannelList[master]:
                            # assigning new master to resampled channels
                            self.set_channel_master(channel, master_channel_name)
                        # merging channel lists
                        self.masterChannelList[master_channel_name] += self.masterChannelList[master]
                        # removing old master
                        self.masterChannelList.pop(master)
                        self.pop(master)
        else:
            warn('no data to be resampled')

    def resample_group(self, sampling, channel, new_master_data=None):
        """ Resamples one channel along with its dataGroup

        Parameters
        ----------------
        sampling : float
            resampling interval

        channel : str
            channel name to be resampled (could the master channel)

        new_master_data : array, optional
            master channel data to be applied to the group identified by channel

        Notes
        --------
        Resampling will convert all channels so be careful for big files
        and memory consumption
        """
        def interpolate(new_x, x, y):
            if y.dtype.kind == 'f':
                return interp(new_x, x, y)
            else:
                idx = searchsorted(x, new_x, side='right')
                idx -= 1
                idx = clip(idx, 0, idx[-1])
                return y[idx]

        master_channel = self.get_channel_master(channel)
        old_master_data = self.get_channel_data(master_channel)
        if new_master_data is None:
            new_master_data = arange(old_master_data[0], old_master_data[-1], sampling)
        for Name in list(self.masterChannelList[master_channel]):
            # forces list() because masterChannelList is dynamic, channels can be removed
            if Name == master_channel:  # master channel
                self.set_channel_data(Name, new_master_data)
            else:
                channel_data = self.get_channel_data(Name)
                if channel_data.dtype.kind not in ('S', 'U', 'V'):
                    # if channel not array of string
                    try:
                        self.set_channel_data(Name, interpolate(new_master_data, old_master_data, channel_data))
                    except:
                        if not all(diff(old_master_data) > 0):
                            warn('{} has non regularly increasing master channel {}.\n'
                                 ' Faulty samples will be dropped in related data group'.
                                 format(Name, master_channel))
                            self._clean_uneven_master_data(master_channel)
                            self.set_channel_data(Name, interpolate(new_master_data, old_master_data, channel_data))
                        elif old_master_data is not None and len(old_master_data) != len(channel_data):
                            warn('{} and master channel {} do not have same length'.
                                 format(Name, master_channel))
                    self.remove_channel_conversion(Name)
                else:  # can not interpolate strings, remove channel containing string
                    self.remove_channel(Name)

    def _clean_uneven_master_data(self, master_channel_name):
        """ clean data group having non evenly increasing master

            Parameters
            ----------------
            master_channel_name : str
                master channel name referring to data group to be cleaned

        """
        # create mask
        mask = diff(self.get_channel_data(master_channel_name)) > 0
        # remove samples from data group defined by mask
        for channel in self.masterChannelList[master_channel_name]:
            data = self.get_channel_data(channel).view(MaskedArray)
            data.mask = mask
            self.set_channel_data(channel, data.compressed())

    def cut(self, master_channel, begin=None, end=None):
        """ Cut data

        Parameters
        ----------------
        master_channel : str
            channel to cut data (can be master channel)

        begin : float
            beginning value in master channel from which to start cutting in all channels

        end : float
            ending value in master channel from which to start cutting in all channels

        Notes
        ------
        Only the data groups with same master type as master_channel will be cut (only for mdf4)

        """
        if begin is None and end is None:
            raise Exception('Please input at least one beginning or ending value to cut data')

        master_channel = self.get_channel_master(master_channel)
        master_channel_type = self.get_channel_master_type(master_channel)
        for master in self.masterChannelList:  # for each channel group
            # find corresponding indexes to cut
            master_data = self.get_channel_data(master)
            if master_data is not None and len(master_data) > 0 and \
                    self.get_channel_master_type(master) == master_channel_type:
                # not empty data and same master type
                if begin is not None:
                    start_index = searchsorted(master_data, begin, side='left')
                else:
                    start_index = 0
                if end is not None:
                    end_index = searchsorted(master_data, end, side='right')
                else:
                    end_index = len(master_data)
                if start_index == end_index:
                    # empty array
                    for channel in self.masterChannelList[master]:
                        self.set_channel_data(channel, array([]))
                else:
                    for channel in self.masterChannelList[master]:
                        data = self.get_channel_data(channel)
                        self.set_channel_data(channel, data[start_index: end_index])

    def export_to_csv(self, file_name=None, sampling=None):
        """ Exports mdf data into CSV file

        Parameters
        ----------------
        file_name : str, optional
            file name. If no name defined, it will use original mdf name and path

        sampling : float, optional
            sampling interval. None by default

        Notes
        --------
        Data saved in CSV file be automatically resampled as it is difficult to save in this format
        data not sharing same master channel -> not applicable for mdf4 in case there are master channels
         with various types
        Warning: this can be slow for big data, CSV is text format after all

        """

        if self:  # data in mdf
            import csv
            self.resample(sampling)
            if file_name is None:
                file_name = splitext(self.fileName)[0]
                file_name = file_name + '.csv'
            if self.MDFVersionNumber >= 400:
                encoding = 'utf8'  # mdf4 encoding is unicode
            else:
                encoding = 'latin-1'  # mdf3 encoding is latin-1
            # writes header
            f = open(file_name, "wt", encoding=encoding)
            writer = csv.writer(f, dialect=csv.excel)
            writer.writerow([name for name in self
                             if self.get_channel_data(name).dtype.kind not in ('S', 'U', 'V')
                             and self.get_channel_data(name).ndim <= 1])  # writes channel names
            # writes units
            writer.writerow([self.get_channel_unit(name)
                             for name in self
                             if self.get_channel_data(name).dtype.kind not in ('S', 'U', 'V')
                             and self.get_channel_data(name).ndim <= 1])  # writes units
            # concatenate all channels
            temp = []
            for name in self:
                data = self.get_channel_data(name)
                if data.dtype.kind not in ('S', 'U', 'V') \
                        and data.ndim <= 1:
                    temp.append(data.transpose())
            if temp:
                buf = vstack(temp)
                buf = buf.transpose()
                # Write all rows
                r, c = buf.shape
                writer.writerows([list(buf[i, :]) for i in range(r)])
            f.close()
        else:
            warn('no data to be exported')

    def export_to_NetCDF(self, file_name=None, sampling=None):
        """Exports mdf data into netcdf file

        Parameters
        ----------------
        file_name : str, optional
            file name. If no name defined, it will use original mdf name and path

        sampling : float, optional
            sampling interval

        Notes
        --------
        Dependency: scipy
        """
        try:
            from scipy.io import netcdf
        except ImportError:
            warn('scipy.io module not found')
            return

        def clean_name(name):
            allowed_str = ' ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-+_.@'
            buf = ''
            for c in name:
                if c in allowed_str:
                    buf += c
            return buf

        def set_attribute(f, name, value):
            if value is not None and len(value) > 0:  # netcdf does not allow empty strings...
                if value is dict and 'name' in value:
                    value = value['name']
                if value is bytes:
                    value = value.encode('utf-8', 'ignore')
                value = clean_name(value)
                setattr(f, name, value)
            else:
                pass
        if sampling is not None:
            self.resample(sampling)
        if file_name is None:
            file_name = splitext(self.fileName)[0]
            file_name = file_name + '.nc'
        f = netcdf.netcdf_file(file_name, 'w')
        setattr(f, 'Time', self.fileMetadata['time'])
        set_attribute(f, 'Author', self.fileMetadata['author'])
        set_attribute(f, 'Organization', self.fileMetadata['organisation'])
        set_attribute(f, 'ProjectName', self.fileMetadata['project'])
        set_attribute(f, 'Subject', self.fileMetadata['subject'])
        set_attribute(f, 'Comment', self.fileMetadata['comment'])
        # Create dimensions having name of all time channels
        for master in self.masterChannelList:
            f.createDimension(master, len(self.get_channel_data(self.masterChannelList[master][0])))
        # Create variables definition, dimension and attributes
        var = {}
        master_channel_set = set(self.masterChannelList.keys())
        for name in self:
            data = self.get_channel_data(name)
            if data.dtype == 'float64':
                data_type = 'd'
            elif data.dtype == 'float32':
                data_type = 'f'
            elif data.dtype in ['int8', 'int16', 'uint8', 'uint16']:
                data_type = 'h'
            elif data.dtype in ['int32', 'uint32']:
                data_type = 'i'
            elif data.dtype.kind in ['S', 'U']:
                data_type = 'c'
            else:
                data_type = None
                warn('Can not process numpy type {} of channel {}'.format(data.dtype, name))
            if data_type is not None:
                # create variable
                cleaned_name = clean_name(name)
                if len(self.masterChannelList) == 1:  # mdf resampled
                    var[name] = f.createVariable(cleaned_name, data_type, (list(self.masterChannelList.keys())[0],))
                else:  # not resampled
                    var[name] = f.createVariable(cleaned_name, data_type, (self.get_channel_master(name),))
                # Create attributes
                set_attribute(var[name], 'title', cleaned_name)
                set_attribute(var[name], 'units', self.get_channel_unit(name))
                set_attribute(var[name], 'Description', self.get_channel_desc(name))
                if name in master_channel_set:
                    set_attribute(var[name], 'Type', 'Master Channel')
                    set_attribute(var[name], 'datatype', 'master')
                else:
                    set_attribute(var[name], 'Type', 'Data Channel')
        # put data in variables
        for name in self:
            var[name] = self.get_channel_data(name)
        f.close()

    def export_to_hdf5(self, file_name=None, sampling=None, compression=None, compression_opts=None):
        """Exports mdf class data structure into hdf5 file

        Parameters
        ----------------
        file_name : str, optional
            file name. If no name defined, it will use original mdf name and path

        sampling : float, optional
            sampling interval.

        compression : str, optional
            HDF5 compression algorithm. Valid options are 'gzip', 'lzf'.
            gzip compression recommended for portability.
            szip compression not supported due to legal reasons.

        compression_opts : int, optional
            HDF5 gzip compression level, 0-9. Only valid if gzip compression is used.
            Level 4 (default) recommended for best balance between compression and time.

        Notes
        --------
        The maximum attributes will be stored.
        Data structure will be similar has it is in masterChannelList attribute.
        Dependency: h5py
        """
        #
        try:
            import h5py
            import os
        except ImportError:
            warn('h5py not found')
            return

        def set_attribute(obj, name, value):
            if value is not None and len(value) > 0:
                try:
                    if value is dict and 'name' in value:
                        value = value['name']
                    obj.attrs[name] = value
                except:
                    pass
            else:
                pass

        if sampling is not None:
            self.resample(sampling)
        if file_name is None:
            file_name = splitext(self.fileName)[0]
            file_name = file_name + '.hdf'
        if compression is not None:
            compression = compression.lower()
            if compression not in ['gzip', 'lzf']:
                compression = None
                compression_opts = None
            elif compression == 'lzf':
                compression_opts = None
            if compression_opts not in range(10):
                compression_opts = None
        else:
            compression_opts = None

        f = h5py.File(file_name, 'w')  # create hdf5 file
        # create group in root associated to file
        file_group = f.create_group(os.path.basename(file_name))
        file_group.attrs['Time'] = self.fileMetadata['time']
        set_attribute(file_group, 'Author', self.fileMetadata['author'])
        set_attribute(file_group, 'Organization', self.fileMetadata['organisation'])
        set_attribute(file_group, 'ProjectName', self.fileMetadata['project'])
        set_attribute(file_group, 'Subject', self.fileMetadata['subject'])
        set_attribute(file_group, 'Comment', self.fileMetadata['comment'])
        master_type_dict = {0: 'None', 1: 'Time', 2: 'Angle', 3: 'Distance', 4: 'Index', None: 'None'}
        if len(self.masterChannelList) > 1:
            # if several time groups of channels, not resampled
            groups = {}
            n_groups = 0
            grp = {}
            for channel in self:
                channel_data = self.get_channel_data(channel)
                master_name = self.get_channel_master(channel)
                if masterField in self[channel] and master_name not in groups:
                    # create new data group
                    n_groups += 1
                    if master_name != '' \
                            and master_name is not None:
                        group_name = master_name
                    else:
                        group_name = masterField + str(n_groups)
                    groups[group_name] = n_groups
                    grp[n_groups] = file_group.create_group(group_name)
                    set_attribute(grp[n_groups], masterField, master_name)
                    set_attribute(grp[n_groups], masterTypeField,
                                  master_type_dict[self.get_channel_master_type(channel)])
                elif masterField in self[channel] and master_name in groups:
                    group_name = master_name
                if channel_data.dtype.kind not in ('U', 'O'):  # not supported type
                    channel_name = _convert_to_hdf5_name(channel)
                    dset = grp[groups[group_name]].create_dataset(channel_name,
                                                                  data=channel_data,
                                                                  compression=compression,
                                                                  compression_opts=compression_opts,
                                                                  chunks=True)
                    set_attribute(dset, unitField, self.get_channel_unit(channel))
                    if descriptionField in self[channel]:
                        set_attribute(dset, descriptionField, self.get_channel_desc(channel))
        else:  # resampled or only one time for all channels : no groups
            master_name = list(self.masterChannelList.keys())[0]
            set_attribute(file_group, masterField, master_name)
            set_attribute(file_group, masterTypeField,
                          master_type_dict[self.get_channel_master_type(master_name)])
            for channel in self:
                channel_data = self.get_channel_data(channel)
                if channel_data.dtype.kind not in ('U', 'O'):  # not supported type
                    channel_name = _convert_to_hdf5_name(channel)
                    dset = file_group.create_dataset(channel_name, data=channel_data,
                                                     compression=compression,
                                                     compression_opts=compression_opts,
                                                     chunks=True)
                    set_attribute(dset, unitField, self.get_channel_unit(channel))
                    if descriptionField in self[channel]:
                        set_attribute(dset, descriptionField, self.get_channel_desc(channel))
        f.close()

    def export_to_matlab(self, file_name=None):
        """Export mdf data into Matlab file preferrably in format 7.3

        Parameters
        ----------------
        file_name : str, optional
            file name. If no name defined, it will use original mdf name and path

        Notes
        --------
        This method will dump all data into Matlab file but you will loose below information:
        - unit and descriptions of channel
        - data structure, what is corresponding master channel to a channel.
        Channels might have then different lengths.
        Dependency: hdf5storage, scipy
        """
        # export class data structure into .mat file
        try:
            from hdf5storage import savemat
        except ImportError:
            warn('hdf5storage module not found')
            try:
                from scipy.io import savemat
            except ImportError:
                warn('scipy also module not found')
                return
        if file_name is None:
            file_name = splitext(self.fileName)[0]
            file_name = file_name + '.mat'
        # convert self into simple dict without and metadata
        temp = {'masterChannelList': {}}
        for master in self.masterChannelList:
            temp['masterChannelList'][master] = {}
        for channel in self:
            data = self.get_channel_data(channel)
            if data.dtype.kind not in ('S', 'U', 'V'):  # does not like special characters chains, skip
                channel_name = _convert_to_matlab_name(channel)
                if len(channel_name) > 0 and channel_name is not None:  # adds description and unit if existing
                    temp[channel_name] = data
                    if data.dtype.base.name == 'float16':
                        temp[channel_name] = data.astype(float32)
                    if isinstance(data, MaskedArray):
                        temp[channel_name] = data.filled()
                    unit = self.get_channel_unit(channel)
                    desc = self.get_channel_desc(channel)
                    master = self.get_channel_master(channel)
                    if unit or desc or master == channel:
                        temp['masterChannelList'][master][channel_name] = {}
                        if desc:
                            temp['masterChannelList'][master][channel_name]['description'] = desc
                        if unit:
                            temp['masterChannelList'][master][channel_name]['unit'] = unit
                        if master == channel:
                            temp['masterChannelList'][master][channel_name]['masterType'] = \
                                self.get_channel_master_type(channel)
                elif channel_name is not None:
                    warn(u'Could not export {}, name is not compatible with Matlab'.format(channel))
        try:
            savemat(file_name, temp, format='7.3', long_field_names=True, oned_as='column',
                    structured_numpy_ndarray_as_struct=True)
        except:
            savemat(file_name, temp, long_field_names=True, format='5')

    def export_to_excel(self, file_name=None):
        """Exports mdf data into excel 95 to 2003 file

        Parameters
        ----------------
        file_name : str, optional
            file name. If no name defined, it will use original mdf name and path

        Notes
        --------
        xlwt is not fast even for small files, consider other binary formats like HDF5 or Matlab.
        If there are more than 256 channels, data will be saved over different worksheets.
        Also Excel 2003 is becoming rare these days, prefer using exportToXlsx.
        Dependencies: xlwt for python 2.6+, xlwt3 for python 3.2+
        """
        try:
            import xlwt3 as xlwt
        except ImportError:
            warn('xlwt3 module missing')
            return
        if file_name is None:
            file_name = splitext(self.fileName)[0]
            file_name = file_name + '.xls'
        style_text = xlwt.easyxf('font: name Times New Roman, color-index black, bold off')
        coding = 'utf-8'
        wb = xlwt.Workbook(encoding=coding)
        channel_list = list(self.keys())
        units = [self.get_channel_unit(channel) for channel in self]
        # Excel 2003 limits
        max_cols = 255
        max_lines = 65535
        workbook_number = int(ceil(len(channel_list) * 1.0 / (max_cols * 1.0)))
        too_long_channels = []
        # split columns in several worksheets if more than 256 cols
        for workbook in range(workbook_number):
            ws = wb.add_sheet('Sheet' + str(workbook))  # , cell_overwrite_ok = True )
            if workbook == workbook_number - 1:  # last sheet
                column_range = list(range(workbook * max_cols, len(channel_list)))
            elif workbook < workbook_number - 1 and workbook_number > 1:  # first sheets
                column_range = list(range(workbook * max_cols, (workbook + 1) * max_cols))
            for col in column_range:
                # write header
                ws.write(0, col - workbook * max_cols)
                ws.write(1, col - workbook * max_cols)
                vector = self.get_channel_data(channel_list[col])  # data vector
                if not len(vector) > max_lines:
                    if vector.dtype.kind not in ['S', 'U']:  # if not a string or unicode
                        [ws.row(row + 2).set_cell_number(col - workbook * max_cols, vector[row])
                         for row in list(range(len(vector)))]
                    else:  # it's a string, cannot write for the moment
                        [ws.row(row + 2).set_cell_text(col - workbook * max_cols, vector[row])
                         for row in list(range(len(vector)))]
                else:  # channel too long, written until max Excel line limit
                    if vector.dtype.kind not in ['S', 'U']:  # if not a string
                        [ws.row(row + 2).set_cell_number(col - workbook * max_cols, vector[row])
                         for row in list(range(max_lines))]
                    else:  # it's a string, cannot write for the moment
                        [ws.row(row + 2).set_cell_text(col - workbook * max_cols, vector[row])
                         for row in list(range(max_lines))]
                    # to later warn user the channel is not completely written
                    too_long_channels.append(channel_list[col])
        wb.save(file_name)  # writes workbook on HDD
        if len(too_long_channels) > 0:  # if not empty, some channels have been not processed
            warn('Following channels were too long to be processed completely,'
                 ' maybe you should resample : {}'.format(too_long_channels))

    def export_to_xlsx(self, file_name=None):
        """Exports mdf data into excel 2007 and 2010 file

        Parameters
        ----------------
        file_name : str, optional
            file name. If no name defined, it will use original mdf name and path

        Notes
        --------
        It is recommended to export resampled data for performances
        Dependency: openpyxl
        """
        try:
            import openpyxl
            from openpyxl.utils.exceptions import IllegalCharacterError
        except ImportError:
            warn('Module openpyxl missing')
            return
        if file_name is None:
            file_name = splitext(self.fileName)[0]
            file_name = file_name + '.xlsx'

        warn('Creating Excel sheet')
        wb = openpyxl.workbook.Workbook()
        ws = wb.active  # get_active_sheet()

        n_col = 1
        for master in self.masterChannelList:
            channels = self.masterChannelList[master]
            for channel in channels:
                data = self.get_channel_data(channel)
                if data.ndim <= 1:  # not an array channel
                    ws.cell(row=1, column=n_col).value = channel
                    ws.cell(row=2, column=n_col).value = self.get_channel_unit(channel)
                    try:
                        if data.ndim <= 1:  # not an array channel
                            if data.dtype.kind in ('i', 'u') or 'f4' in data.dtype.str:
                                for r, cell_data in enumerate(data):
                                    ws.cell(row=r + 3, column=n_col).value = float64(cell_data)
                            elif not data.dtype.kind == 'V':
                                for r, cell_data in enumerate(data):
                                    ws.cell(row=r + 3, column=n_col).value = cell_data
                    except IllegalCharacterError:
                        warn(u'could not export {}'.format(channel))
                    n_col += 1

        warn('Writing file, please wait')
        wb.save(file_name)

    def keep_channels(self, channel_set):
        """ keeps only a list or set of channels and removes the other channels

        Parameters
        ----------------
        channel_set : list or set of str
            list or set of channel names
        """
        if isinstance(channel_set, list):
            channel_set = set(channel_set)
        remove_channels = []
        master_channel_set = set(self.masterChannelList.keys())
        for channel in self:
            if channel not in channel_set and masterField not in channel \
                    and channel not in master_channel_set:
                # avoid to remove master channels otherwise problems with resample
                remove_channels.append(channel)
        if not len(remove_channels) == 0:
            [self.masterChannelList[self.get_channel_master(channel)].remove(channel) for channel in remove_channels]
            [self.pop(channel) for channel in remove_channels]

    def concat_mdf(self, mdf_class):
        """Concatenate data of input mdf class after the current one.

        Parameters
        ----------------
        mdf_class : Mdf
            mdf class instance to be concatenated after self

        Notes
        --------
        It creates union of both channel lists and fills with Nan for unknown sections in channels
        If one channel is not present in both classes, masked array is created
        If invalid bytes are present, masked array are created
        """
        first_class_masters = set(self.masterChannelList.keys())
        second_class_masters = set(mdf_class.masterChannelList.keys())
        union_masters = first_class_masters | second_class_masters
        # find max time of all masters from both classes
        first_masters = {}
        for master_channel_name in first_class_masters:
            type = self.get_channel_master_type(master_channel_name)
            data = self.get_channel_data(master_channel_name)
            if type not in first_masters:
                first_masters[type] = {}
                first_masters[type]['max'] = data[-1]
            # first_masters[type]['sampling'] = mean(diff(data))  # sampling
            first_masters[type]['max'] = max([data[-1], first_masters[type]['max']])
        second_masters = {}
        for master_channel_name in second_class_masters:
            type = mdf_class.get_channel_master_type(master_channel_name)
            data = mdf_class.get_channel_data(master_channel_name)
            if type not in second_masters:
                second_masters[type] = {}
                second_masters[type]['max'] = data[-1]
            # second_masters[type]['sampling'] = mean(diff(data))  # sampling
            second_masters[type]['max'] = max([data[-1], second_masters[type]['max']])

        for master_channel_name in union_masters:
            if master_channel_name in first_class_masters and master_channel_name in second_class_masters:
                # same master name in both classes
                first_class_length = len(self.get_channel_data(master_channel_name))
                first_class_channels = set(self.masterChannelList[master_channel_name])
                second_class_channels = set(mdf_class.masterChannelList[master_channel_name])
                first_class_group_number = self[master_channel_name][idField][0][0]
                second_class_group_number = mdf_class[master_channel_name][idField][0][0]
                if self.MDFVersionNumber >= 400:
                    invalid_channel = 'invalid_bytes{}'.format(first_class_group_number)
                    if invalid_channel in first_class_channels:
                        # invalid bits present, converting to masked array
                        # but invalid bit channels should be removed, updating channel sets
                        self.apply_all_invalid_bit()
                        first_class_channels.remove(invalid_channel)
                if mdf_class.MDFVersionNumber >= 400:
                    invalid_channel = 'invalid_bytes{}'.format(second_class_group_number)
                    if invalid_channel in second_class_channels:
                        mdf_class.apply_all_invalid_bit()
                        second_class_channels.remove(invalid_channel)
                unioned_set = first_class_channels | second_class_channels
                second_class_length = len(mdf_class.get_channel_data(master_channel_name))
                total_length = first_class_length + second_class_length
                # remove master channel
                unioned_set.remove(master_channel_name)
                # merge master channels
                data = self.get_channel_data(master_channel_name)
                mdf_data = mdf_class.get_channel_data(master_channel_name)
                temp = empty(total_length, dtype=data.dtype)
                temp[:first_class_length] = data
                if self.get_channel_master_type(master_channel_name) == 1:
                    # master of type time
                    offset = mean(diff(mdf_data))  # sampling
                    offset = data[-1] + offset  # offset
                    temp[first_class_length:] = mdf_data + offset
                else:
                    temp[first_class_length:] = mdf_data
                self.set_channel_data(master_channel_name, temp)
                for channel in unioned_set:
                    if channel in first_class_channels and channel in second_class_channels:
                        # channel exists in both classes
                        data = self.get_channel_data(channel)
                        if isinstance(data, MaskedArray):
                            temp = ma_empty(total_length, dtype=data.dtype)
                        else:
                            temp = empty(total_length, dtype=data.dtype)
                        temp[:first_class_length] = data
                        temp[first_class_length:] = mdf_class.get_channel_data(channel)
                        self.set_channel_data(channel, temp)
                    elif channel in second_class_channels:
                        self[channel] = mdf_class[channel]  # initialise all fields, units, descriptions, etc.
                        # new channel for self from mdfClass
                        data = self.get_channel_data(channel)
                        if isinstance(data, MaskedArray):
                            temp = ma_empty(total_length, dtype=data.dtype)
                        else:
                            temp = empty(total_length, dtype=data.dtype)
                        temp[:first_class_length] = nan  # fill with NANs
                        temp[first_class_length:] = data
                        self.set_channel_data(channel, temp)
                    else:  # channel missing in mdfClass
                        data = self.get_channel_data(channel)
                        if isinstance(data, MaskedArray):
                            temp = ma_empty(total_length, dtype=data.dtype)
                        else:
                            temp = empty(total_length, dtype=data.dtype)
                        temp[:first_class_length] = data
                        temp[first_class_length:] = nan  # fill with NANs
                        self.set_channel_data(channel, temp)
            elif master_channel_name in first_class_masters:
                first_class_channels = set(self.masterChannelList[master_channel_name])
                first_class_group_number = self[master_channel_name][idField][0][0]
                if self.MDFVersionNumber >= 400:
                    invalid_channel = 'invalid_bytes{}'.format(first_class_group_number)
                    if invalid_channel in first_class_channels:
                        # invalid bits present, converting to masked array
                        # but invalid bit channels has been removed, updating channel sets
                        self.apply_all_invalid_bit()
                        first_class_channels.remove(invalid_channel)
                master_type = self.get_channel_master_type(master_channel_name)
                if master_type in second_masters:
                    # same kind of master existing in both classes
                    master_data = self.get_channel_data(master_channel_name)
                    first_class_length = len(master_data)
                    first_master_end = master_data[-1]
                    total_length = first_class_length + 1
                    # extend by 1 sample all channels and convert to masked array
                    for channel in first_class_channels:
                        data = self.get_channel_data(channel)
                        temp = ma_empty(total_length, dtype=data.dtype)
                        temp[:first_class_length] = data.view(MaskedArray)
                        temp[-1] = masked  # mask last sample
                        self.set_channel_data(channel, temp)
                    master_data = self.get_channel_data(master_channel_name)
                    # last master channel value adjusted and becomes valid
                    master_data[-1] = first_master_end + second_masters[master_type]['max']
                    self.set_channel_data(master_channel_name, master_data)
            else:
                second_class_channels = set(mdf_class.masterChannelList[master_channel_name])
                second_class_group_number = mdf_class[master_channel_name][idField][0][0]
                if mdf_class.MDFVersionNumber >= 400:
                    invalid_channel = 'invalid_bytes{}'.format(second_class_group_number)
                    if invalid_channel in second_class_channels:
                        mdf_class.apply_all_invalid_bit()
                        second_class_channels.remove(invalid_channel)
                master_type = mdf_class.get_channel_master_type(master_channel_name)
                if master_type in first_masters:
                    # same kind of master existing in both classes
                    master_data = mdf_class.get_channel_data(master_channel_name)
                    second_class_length = len(master_data)
                    second_master_end = master_data[-1]
                    total_length = second_class_length + 1
                    for channel in second_class_channels:
                        self[channel] = mdf_class[channel]
                        data = self.get_channel_data(channel)
                        temp = ma_empty(total_length, dtype=data.dtype)
                        temp[1:] = data
                        temp[0] = masked  # mask first sample
                        self.set_channel_data(channel, temp)
                    master_data = self.get_channel_data(master_channel_name)
                    # first master channel value adjusted and becomes valid
                    master_data += first_masters[master_type]['max']
                    master_data[0] = 0
                    self.set_channel_data(master_channel_name, master_data)

    def merge_mdf(self, mdf_class):
        """merge data of input mdf class with the current one.

        Parameters
        ----------------
        mdf_class : Mdf
            mdf class instance to be merged with self

        Notes
        --------
        If there are common channel names between the 2 mdf, channels are renamed to make them unique
        """
        # apply eventual invalid bytes, most probably
        if self.MDFVersionNumber >= 400:
            self.apply_all_invalid_bit()
        if mdf_class.MDFVersionNumber >= 400:
            mdf_class.apply_all_invalid_bit()
        # check for same channel names
        first_channels = set(self.keys())
        second_channels = set(mdf_class.keys())
        common_channels = first_channels & second_channels
        for channel in common_channels:
            mdf_class.rename_channel(channel,
                                     '{}_{}'.format(channel,
                                                    mdf_class[channel][idField][0][0]))
        # copy the data
        for channel in mdf_class:
            self[channel] = mdf_class[channel]
        # merge the 2 masterChannelList
        masterChannelList = {**self.masterChannelList, **mdf_class.masterChannelList}
        self.masterChannelList = masterChannelList

    def convert_to_pandas(self, sampling=None):
        """converts mdf data structure into pandas dataframe(s)

        Parameters
        ----------------
        sampling : float, optional
            resampling interval

        Notes
        --------
        One pandas dataframe is converted per data group (one master per data group)
        """
        # convert data structure into pandas module
        try:
            import pandas as pd
        except ImportError:
            warn('Module pandas missing')
            return
        if sampling is not None:
            self.resample(sampling)
        for group in self.masterChannelList:
            self[group + '_group'] = self.return_pandas_dataframe(group)
            # clean rest of self from data and time channel information
            [self[channel].pop(dataField) for channel in self.masterChannelList[group]]
            [self[channel].pop(masterField) for channel in self.masterChannelList[group] if
             masterField in self[channel]]
        self.masterGroups = []  # save time groups name in list
        [self.masterGroups.append(group + '_group') for group in self.masterChannelList]
        self.masterChannelList = {}
        self._pandasframe = True

    def return_pandas_dataframe(self, master_channel_name):
        """returns a dataframe of a raster described by its master channel name

        Parameters
        ----------------
        master_channel_name : str
            master channel name, key to a raster to be returned as pandas dataframe

        Return
        ---------
        pandas dataframe of raster or data group
        """
        try:
            import pandas as pd
        except ImportError:
            warn('Module pandas missing')
            return
        if master_channel_name in self:
            if master_channel_name not in self.masterChannelList[master_channel_name]:
                warn('no master channel in group {}'.format(master_channel_name))
            if master_channel_name in self.masterChannelList[master_channel_name]:
                if self.get_channel_master_type(master_channel_name) == 1:
                    # master channel exists and is time type
                    # convert time channel into timedelta
                    datetime_info = datetime64(int(self.fileMetadata['time'] * 1E9), 'ns')
                    time = datetime_info + array(self.get_channel_data(master_channel_name) * 1E9,
                                                 dtype='timedelta64[ns]')
                    temporary_dataframe = pd.DataFrame(index=time)
                else:  # not time master channel
                    temporary_dataframe = pd.DataFrame(index=self.get_channel_data(master_channel_name))
            else:  # no master channel
                temporary_dataframe = pd.DataFrame()
            channel_dict = {key: None for key in self.masterChannelList[master_channel_name]}
            for key in channel_dict.keys():
                data = self.get_channel_data(key)
                if data.dtype.byteorder not in ['=', '|']:
                    data = data.byteswap().newbyteorder()
                if data.ndim == 1 and data.shape[0] == temporary_dataframe.shape[0] \
                        and not data.dtype.char == 'V':
                    channel_dict[key] = data
            temporary_dataframe = pd.DataFrame(data=channel_dict, index=temporary_dataframe.index)
            return temporary_dataframe
        else:
            warn('Master channel name not in mdf')
            return

    def export_to_parquet(self, file_name=None):
        """Exports mdf data into parquet file

        Parameters
        ----------------
        file_name : str, optional
            file name. If no name defined, it will use original mdf name and path with .parquet extension
        """
        try:
            from fastparquet import write as write_parquet
        except ImportError:
            warn('fastparquet not installed')
            return
        if file_name is None:
            file_name = splitext(self.fileName)[0]
            file_name = file_name + '.parquet'
        for master_channel_name in self.masterChannelList:
            frame = self.return_pandas_dataframe(master_channel_name)
            if frame is not None:
                write_parquet(file_name, frame, compression='GZIP')


if __name__ == "__main__":
    if osname == 'nt':
        from multiprocessing import freeze_support
        freeze_support()
    parser = ArgumentParser(prog='mdfreader', description='reads mdf file')
    parser.add_argument('--export', dest='export', default=None,
                        choices=['CSV', 'HDF5', 'Matlab', 'Xlsx', 'Excel', 'NetCDF', 'MDF', 'parquet'],
                        help='Export after parsing to defined file type')
    parser.add_argument('--list_channels', dest='list_channels', action='store_true',
                        help='list of channels in file')
    parser.add_argument('fileName', help='mdf file name')
    parser.add_argument('--channelList', dest='channelList', nargs='+', type=str,
                        default=None, help='list of channels to only read')
    parser.add_argument('--plot', dest='plot_channel_list', nargs='+', type=str,
                        default=None, help='plots list of channels')
    parser.add_argument('--noConversion', action='store_false',
                        help='Do not convert raw channel data \
            to physical values just after reading. Useful if you have memory concerns')
    parser.add_argument('--filterChannelNames', action='store_true',
                        help='activates channel name filtering; \
            removes modules names separated by a point character')
    parser.add_argument('--noDataLoading', action='store_true',
                        help='Do not load data in memory, creates dummy \
            mdf structure and load data from file when needed')
    parser.add_argument('--compression', action='store_true',
                        help='compress data in memory')

    args = parser.parse_args()

    temp = Mdf(file_name=args.fileName, channel_list=args.channelList,
               convert_after_read=args.convertAfterRead,
               filter_channel_names=args.filterChannelNames,
               no_data_loading=args.noDataLoading,
               compression=args.compression)
    if args.export is not None:
        if args.export == 'CSV':
            temp.export_to_csv()
        elif args.export == 'HDF5':
            temp.export_to_hdf5()
        elif args.export == 'Matlab':
            temp.export_to_matlab()
        elif args.export == 'Xlsx':
            temp.export_to_xlsx()
        elif args.export == 'Excel':
            temp.export_to_excel()
        elif args.export == 'NetCDF':
            temp.export_to_NetCDF()
        elif args.export == 'MDF':
            temp.write()
        elif args.export == 'parquet':
            temp.export_to_parquet()
    if args.plot_channel_list is not None:
        temp.plot(args.plot_channel_list)
    if args.list_channels:
        print(temp.masterChannelList)
